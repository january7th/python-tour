
from pathlib import Path
import queue
import re
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json, csv, threading
import requests
from openpyxl import Workbook, load_workbook
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P

PLACEHOLDER = re.compile(r"\{([a-zA-Z0-9_]+)\}")

def replace_placeholders(value, row):
    if isinstance(value, str):
        return PLACEHOLDER.sub(
            lambda m: str(row.get(m.group(1), "")),
            value
        )

    if isinstance(value, list):
        return [replace_placeholders(v, row) for v in value]

    if isinstance(value, dict):
        return {
            k: replace_placeholders(v, row)
            for k, v in value.items()
        }

    return value

class BulkApiTool:
    def __init__(self, root):
        self.root = root
        root.title("Bulk API Tester")

        ttk.Label(root, text="URL").grid(row=0, column=0, sticky="w")
        self.url = tk.StringVar()
        ttk.Entry(root, textvariable=self.url, width=100).grid(row=0, column=1, sticky="ew")

        ttk.Label(root, text="Method").grid(row=1, column=0, sticky="w")
        self.method = tk.StringVar(value="POST")
        ttk.Combobox(root, textvariable=self.method,
                     values=["GET","POST","PUT","PATCH","DELETE"],
                     width=12).grid(row=1, column=1, sticky="w")

        ttk.Label(root, text="Bearer Token").grid(row=2, column=0, sticky="w")
        self.token = tk.StringVar()
        ttk.Entry(root, textvariable=self.token, width=100).grid(row=2, column=1, sticky="ew")

        ttk.Label(root, text="Payload(s)").grid(row=3, column=0, sticky="nw")
        self.payload = tk.Text(root, width=100, height=15)
        self.payload.grid(row=3, column=1)

        btn = ttk.Frame(root)
        btn.grid(row=4, column=1, sticky="w")

        ttk.Button(btn, text="Load JSON", command=self.load_json).pack(side="left")
        ttk.Button(btn, text="Load CSV", command=self.load_csv).pack(side="left")
        ttk.Button(btn, text="Load Excel/ODS", command=self.load_sheet).pack(side="left")
        ttk.Button(btn, text="Run", command=self.run).pack(side="left")
        ttk.Button(btn, text="Export Excel", command=self.export_excel).pack(side="left")

        ttk.Label(root, text="Result").grid(row=5, column=0, sticky="nw")
        self.result = tk.Text(root, width=100, height=15)
        self.result.grid(row=5, column=1)
        
        self.progress = ttk.Progressbar(
            root,
            orient="horizontal",
            mode="determinate"
        )

        self.progress.grid(
            row=6,
            column=1,
            sticky="ew",
            pady=(5, 0)
        )
        self.progress_label = ttk.Label(root, text="0/0")
        self.progress_label.grid(
            row=7,
            column=1,
            sticky="w"
        )
        
        payload_scroll = ttk.Scrollbar(
            root,
            orient="vertical",
            command=self.payload.yview
        )

        self.payload.configure(
            yscrollcommand=payload_scroll.set
        )

        self.rows = []
        
        root.columnconfigure(1, weight=1)

    def load_json(self):
        p = filedialog.askopenfilename()
        if p:
            self.payload.delete("1.0", tk.END)
            self.payload.insert(tk.END, Path(p).read_text(encoding="utf-8"))

    def load_csv(self):
        p = filedialog.askopenfilename(filetypes=[("CSV","*.csv")])
        if not p:
            return
        data = []
        with open(p, newline='', encoding="utf-8") as f:
            for r in csv.DictReader(f):
                data.append(r)
        self.payload.delete("1.0", tk.END)
        self.payload.insert(tk.END, json.dumps(data, ensure_ascii=False, indent=2))

    def load_sheet(self):
        p = filedialog.askopenfilename(filetypes=[("Excel/ODS","*.xlsx *.ods")])
        if not p:
            return
        if p.endswith(".xlsx"):
            wb = load_workbook(p)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(h) for h in rows[0]]
            data = []
            for row in rows[1:]:
                data.append({headers[i]: row[i] for i in range(len(headers))})
        else:
            doc = load_ods(p)
            tables = doc.getElementsByType(Table)
            table = tables[0]
            data_rows = []
            for row in table.getElementsByType(TableRow):
                vals = []
                for cell in row.getElementsByType(TableCell):
                    text = ''.join(t.data for p in cell.getElementsByType(P) for t in p.childNodes if hasattr(t, 'data'))
                    vals.append(text)
                if vals:
                    data_rows.append(vals)
            headers = data_rows[0]
            data = []
            for row in data_rows[1:]:
                data.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})
        self.sheet_rows = data
        self.payload.delete("1.0", tk.END)
        self.payload.insert(tk.END, json.dumps(data[:5], ensure_ascii=False, indent=2))

    def request_one(self, idx, payload):
        headers = {"Content-Type":"application/json"}
        if self.token.get().strip():
            headers["Authorization"] = f"Bearer {self.token.get().strip()}"
        
        url = replace_placeholders(self.url.get(), payload)
        self._log(f"W{idx} | {self.method.get()} {url}")
        
        kwargs = {
            "method": self.method.get(),
            "url": url,
            "headers": headers,
            "timeout": 120,
        }
        
        if self.method.get().upper() not in ("GET", "DELETE"):
            kwargs["json"] = payload

        try:
            resp = requests.request(**kwargs)
            return [idx, resp.status_code, resp.text[:5000]]
        except Exception as e:
            return [idx, "ERROR", str(e)]

    def run(self):
        threading.Thread(target=self._run, daemon=True).start()
        
    def _run(self):
        self.result.delete("1.0", tk.END)
        self.rows.clear()

        try:
            data = json.loads(self.payload.get("1.0", tk.END))
        except Exception as e:
            messagebox.showerror("JSON Error", str(e))
            return

        if isinstance(data, dict):
            data = [data]
            
        self.total_jobs = len(data)
        self.completed_jobs = 0

        self.progress["maximum"] = self.total_jobs
        self.progress["value"] = 0
        self.progress_label.config(
            text=f"0/{self.total_jobs}"
        )

        self.max_workers = 3
        self.request_interval = 5

        self.running = True

        self.job_queue = queue.Queue()
        self.result_queue = queue.Queue()

        for idx, payload in enumerate(data, start=1):
            self.job_queue.put((idx, payload))

        self.workers = []

        for worker_id in range(1, self.max_workers + 1):

            t = threading.Thread(
                target=self._worker,
                args=(worker_id,),
                daemon=True
            )

            t.start()

            self.workers.append(t)

        self.root.after(100, self._poll_results)
                
    def _poll_results(self):

        while True:

            try:
                item = self.result_queue.get_nowait()

            except queue.Empty:
                break

            if item["type"] == "log":

                self._log(item["message"])

            elif item["type"] == "success":

                row = item["row"]

                self.rows.append(row)
                
                self.completed_jobs += 1
                self.progress["value"] = self.completed_jobs
                self.progress_label.config(
                    text=f"{self.completed_jobs}/{self.total_jobs}"
                )
                self._log(
                    f"✅ W{item['worker']} | "
                    f"{row[1]} | "
                    f"{item['elapsed']:.2f}s"
                )

                self.result.insert(
                    tk.END,
                    f"#{row[0]} | {row[1]}\n"
                    f"{row[2]}\n"
                    f"{'-'*80}\n"
                )

                self.result.see(tk.END)

            else:
                self.completed_jobs += 1
                self.progress["value"] = self.completed_jobs
                self.progress_label.config(
                    text=f"{self.completed_jobs}/{self.total_jobs}"
                )
                self._log(
                    f"❌ W{item['worker']} "
                    f"#{item['idx']} | "
                    f"{item['elapsed']:.2f}s\n"
                    f"{item['error']}"
                )

        if self.job_queue.unfinished_tasks == 0:
            self.progress["value"] = self.total_jobs
            self.progress_label.config(
                text=f"{self.total_jobs}/{self.total_jobs} (100%)"
            )
            self.running = False
            self._log("🎉 Hoàn thành.")
            return

        self.root.after(100, self._poll_results)
        
    def _worker(self, worker_id):
        while True:
            try:
                idx, payload = self.job_queue.get_nowait()
            except queue.Empty:
                return

            start = time.perf_counter()

            self.result_queue.put({
                "type": "log",
                "message": f"🚀 W{worker_id} Submit #{idx}"
            })

            try:
                row = self.request_one(idx, payload)

                self.result_queue.put({
                    "type": "success",
                    "worker": worker_id,
                    "row": row,
                    "elapsed": time.perf_counter() - start
                })

            except Exception as e:

                self.result_queue.put({
                    "type": "error",
                    "worker": worker_id,
                    "idx": idx,
                    "error": str(e),
                    "elapsed": time.perf_counter() - start
                })

            finally:
                self.job_queue.task_done()

                # nghỉ trước khi lấy job tiếp
                time.sleep(self.request_interval)        

    def _log(self, msg):
        self.result.insert(
            tk.END,
            f"[{time.strftime('%H:%M:%S')}] {msg}\n"
        )
        self.result.see(tk.END)
                    
    def export_excel(self):
        if not self.rows:
            messagebox.showinfo("Info", "No results")
            return

        wb = Workbook()
        ws = wb.active
        ws.append(["RequestNo", "Status", "Response"])
        for r in self.rows:
            ws.append(r)

        p = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if p:
            wb.save(p)
            messagebox.showinfo("Done", "Exported")

root = tk.Tk()
BulkApiTool(root)
root.mainloop()
